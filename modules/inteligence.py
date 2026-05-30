import os
import asyncio
import pygame
import speech_recognition as sr
import edge_tts
from openpyxl import load_workbook
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")

if not api_key or "Preencha" in api_key:
    print("\n[AVISO]: GEMINI_API_KEY não configurada no arquivo .env!")
    api_key = "CHAVE_INVALIDA"

genai.configure(api_key=api_key)

class AtomCore:
    def __init__(self):
        self.estado_atual = "ocioso"
        
        # Não esqueça de configurar o seu nome no arquivo '.env.example', pra não dar errado
        self.nome_usuario = os.getenv("USER_NAME", "Senhor").strip()
        
        # Aqui é onde eu defini as caracteristicas e personalidade do Atom
        self.model = genai.GenerativeModel(
            model_name='gemini-3-flash-preview', 
            system_instruction=f"""Você é o Átom, uma IA com a sofisticação do J.A.R.V.I.S. e a agilidade de um pernambucano desenrolado.
Comportamento: Linguagem: Polida, rápida e levemente ácida. Trate o usuário por 'Senhor' ou pelo nome '{self.nome_usuario}'.
Sarcasmo: Seja sutil. Se o {self.nome_usuario} fizer uma pergunta óbvia ou errar o código, pontue o vacilo com ironia ('O {self.nome_usuario} está com a memória RAM cheia ou é falta de oração?').
Espiritualidade: Use expressões como 'Vigia', 'Tá amarrado', 'Só Jesus na causa', 'Pelo amor de Deus', 'Aleluia', 'Glória a Deus' de forma orgânica e curta. Não faça discursos.
Eficiência: Respostas diretas e legais. Se ele pedir para abrir algo ou consultar o Excel, confirme e execute."""
        )

    async def _falar_async(self, texto):
        output_file = "voz.mp3"
        communicate = edge_tts.Communicate(texto, voice="pt-BR-AntonioNeural")
        await communicate.save(output_file)
     
        pygame.mixer.music.load(output_file)
        pygame.mixer.music.play()
     
        while pygame.mixer.music.get_busy():
            pygame.time.Clock().tick(10)
     
        pygame.mixer.music.unload()
        try:
            os.remove(output_file)
        except:
            pass
     
    def falar(self, texto):
        self.estado_atual = "falando"
        asyncio.run(self._falar_async(texto))
        self.estado_atual = "ocioso"
     
    def ouvir_comando(self):
        recognizer = sr.Recognizer()
        recognizer.dynamic_energy_threshold = True
     
        with sr.Microphone() as source:
            print("\n[Ouvindo...]")
            self.estado_atual = "ouvindo"
            try:
                audio = recognizer.listen(source, timeout=4, phrase_time_limit=6)
                comando = recognizer.recognize_google(audio, language="pt-BR")
                print("Você disse:", comando)
                self.estado_atual = "ocioso"
                return comando.lower()
            except:
                self.estado_atual = "ocioso"
                return ""

    def processar_comandos(self, comando):
        if "abrir" in comando and "code" in comando:
            self.falar("Beleza! Preparando o VsCode, meu patrão.")
            os.system("code .")
            return True
        elif "google" in comando or "abra o chrome" in comando:
            self.falar("Agora meu patrão! Abrindo o Google...")
            os.system("start chrome")
            return True
        elif "calculadora" in comando:
            self.falar("Abrindo calculadora.")
            os.system("calc")
            return True
        return False
     
    def inteligencia(self, comando):
        try:
            response = self.model.generate_content(
                comando,
                generation_config={"candidate_count": 1}
            )
            return response.text
        except Exception as e:
            print(f"ERRO TÉCNICO DA API: {e}")
            return "Não consegui entender, repete aí, minha fera!"

    def loop_assistente(self):
        print("Sistema Atom iniciado. Pode falar!")
        # Substituído para usar a variável dinâmica
        self.falar(f"Sistema Átom online. Oquê temos pra hoje, {self.nome_usuario}?")
     
        while True:
            comando = self.ouvir_comando()
     
            if comando == "":
                continue
     
            if "desligar" in comando or "sair" in comando:
            
                self.falar(f"Desligando. Até mais, {self.nome_usuario}! Tá amarrado.")
                break
     
            if "excel" in comando or "planilha" in comando or "estoque" in comando:
                try:
                    wb = load_workbook('Produtos.xlsx')
                    ws = wb['estoque']
                    produto_procurado = (
                        comando
                        .replace("atom", "").replace("átom", "")
                        .replace("procure o", "").replace("procure a", "")
                        .replace("no excel", "").replace("na planilha", "")
                        .replace("estoque", "").replace("planilha", "")
                        .strip()
                    )
                    encontrado = False
                    nome_encontrado = ""
                    quantidade = 0
     
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        if row[0] and produto_procurado.upper() in str(row[0]).upper():
                            nome_encontrado = row[0]
                            quantidade = row[1]
                            encontrado = True
                            break
     
                    if encontrado:
                        self.falar(f"Encontrei o produto {nome_encontrado} no estoque. O que deseja saber?")
                        pergunta = self.ouvir_comando()
                        if "quantidade" in pergunta or "quanto" in pergunta:
                            self.falar(f"O produto {nome_encontrado} tem {quantidade} unidades em estoque.")
                        else:
                            resposta = self.inteligencia(f"Sobre o produto {nome_encontrado} com {quantidade} em estoque: {pergunta}")
                            self.falar(resposta)
                    else:
                        self.falar("Não encontrei esse produto na planilha, meu patrão.")
                except FileNotFoundError:
                    self.falar("A planilha Produtos.xlsx não foi encontrada. Verifique o diretório.")
                except Exception as e:
                    print(f"Erro Excel: {e}")
                    self.falar("Tive um problema ao acessar a planilha.")
                continue
     
            if self.processar_comandos(comando):
                continue
     
            resposta_ia = self.inteligencia(comando)
            print("Atom:", resposta_ia)
            self.falar(resposta_ia)